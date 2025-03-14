import numpy as np
import pandas as pd
from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, numbers
import uvicorn
from datetime import datetime

# ========= PARAMETERS & HELPER FUNCTIONS =========
# Personal Parameters (will be updated via query)
current_age = 65           # default; will be overwritten by computed value from birthdate
max_age = 95

# Social Security - now using user‐supplied FRA benefit.
# (Remove fixed monthly_benefit_65/70.)
# Instead, compute the adjusted monthly benefit based on claim age.
fra_age = 67  # Assumed FRA

def compute_adjusted_benefit(fra_benefit: float, claim_age: int) -> float:
    """
    Computes the monthly benefit at a given claim_age using standard Social Security rules.
    If claim_age < FRA, benefit is reduced by ~0.556% per month early.
    If claim_age > FRA, benefit increases by ~0.667% per month delayed.
    """
    months_difference = (fra_age - claim_age) * 12
    if claim_age < fra_age:
        # Reduction for early claiming (about 0.556% per month)
        reduction_rate = 0.00556
        factor = 1 - (months_difference * reduction_rate)
    elif claim_age > fra_age:
        # Delayed claiming increase (about 0.667% per month)
        increase_rate = 0.00667
        factor = 1 + ((claim_age - fra_age) * 12 * increase_rate)
    else:
        factor = 1
    return fra_benefit * factor

# CBO Projections
trust_fund_depletion_year = 2033
current_year = 2025
benefit_reduction_factor = 0.75

def get_age_ranges(curr_age_val):
    ages = np.arange(curr_age_val, max_age + 1)
    years = np.arange(current_year, current_year + (max_age - curr_age_val) + 1)
    return ages, years

def calculate_federal_income_tax(taxable_income):
    try:
        brackets = []
        with open('tax_brackets_2024.txt', 'r') as file:
            for line in file:
                if line.strip() and not line.startswith('#'):
                    parts = line.strip().split(',')
                    if len(parts) == 3:
                        brackets.append((float(parts[0]), float(parts[1]), float(parts[2])))
        tax = 0
        for lower, upper, rate in brackets:
            if taxable_income <= lower:
                break
            taxable_in_bracket = min(taxable_income - lower, upper - lower)
            tax += taxable_in_bracket * rate
        return tax
    except FileNotFoundError:
        return taxable_income * 0.24

def adjust_benefit_for_cbo_projections(benefit, age, curr_age):
    year = current_year + (age - curr_age)
    return benefit * benefit_reduction_factor if year >= trust_fund_depletion_year else benefit

def estimate_pre_tax_income_needed(after_tax_target, ss_income):
    pre_tax_income = after_tax_target * 1.3
    for _ in range(5):
        provisional_income = (ss_income * 0.5) + (pre_tax_income - ss_income)
        if provisional_income <= 32000:
            taxable_ss = 0
        elif provisional_income <= 44000:
            taxable_ss = min(ss_income * 0.5, (provisional_income - 32000) * 0.5)
        else:
            taxable_ss = min(ss_income * 0.85, (6000)*0.5 + (provisional_income - 44000) * 0.85)
        taxable_income = max(0, (pre_tax_income - ss_income) + taxable_ss - standard_deduction)
        tax_estimate = calculate_federal_income_tax(taxable_income)
        pre_tax_income = after_tax_target + tax_estimate
    return pre_tax_income, tax_estimate

def calculate_rmd(age, account_balance):
    if age < 73:
        return 0
    try:
        rmd_table = {}
        with open('irs_uniform_lifetime_table.txt', 'r') as file:
            for line in file:
                if line.strip() and not line.startswith('#'):
                    parts = line.strip().split(',')
                    age_val = 120 if parts[0] == "120+" else int(parts[0])
                    rmd_table[age_val] = float(parts[1])
        divisor = rmd_table.get(age, rmd_table.get(120, max(1, 90 - age)))
    except FileNotFoundError:
        divisor = max(1, 90 - age)
    return account_balance / divisor

def run_claim_strategy(ages, years, claim_age, current_benefit, curr_age):
    """
    Runs the simulation using monthly amounts.
    current_benefit is the adjusted monthly Social Security benefit for a given model.
    """
    # Here, instead of annual, we work with monthly amounts.
    benefit = current_benefit  
    withdrawals_401k = np.zeros(len(ages))
    withdrawals_non_retirement = np.zeros(len(ages))
    cumulative = np.zeros(len(ages))
    portfolio = np.zeros(len(ages))
    income_taxes_paid = np.zeros(len(ages))
    portfolio[0] = initial_401k
    non_retirement_savings = np.zeros(len(ages))
    non_retirement_savings[0] = other_non_retirement_savings
    after_tax_target = target_income

    for i in range(len(ages)):
        age = ages[i]
        if age >= claim_age:
            adj_benefit = adjust_benefit_for_cbo_projections(benefit, age, curr_age)
            cumulative[i] = adj_benefit if i == 0 else cumulative[i-1] + adj_benefit
        if i > 0:
            ss_income = adjust_benefit_for_cbo_projections(benefit, age, curr_age) if age >= claim_age else 0
            pre_tax_income, tax_est = estimate_pre_tax_income_needed(after_tax_target, ss_income)
            income_taxes_paid[i] = tax_est
            income_need = max(0, pre_tax_income - ss_income)
            real_portfolio_value = portfolio[i-1]
            nominal_portfolio_value = real_portfolio_value * ((1 + inflation_rate) ** i)
            rmd_real = calculate_rmd(age, nominal_portfolio_value) / ((1 + inflation_rate) ** i)
            withdrawals_401k[i] = rmd_real
            excess_deposit = 0
            if rmd_real > income_need:
                excess_deposit = rmd_real - income_need
                remaining_need = 0
            else:
                remaining_need = income_need - rmd_real
                if remaining_need > 0 and portfolio[i-1] > rmd_real:
                    add_from_401k = min(remaining_need, portfolio[i-1] - rmd_real)
                    withdrawals_401k[i] += add_from_401k
                    remaining_need -= add_from_401k
                if remaining_need > 0 and non_retirement_savings[i-1] > 0:
                    from_non_ret = min(remaining_need, non_retirement_savings[i-1])
                    withdrawals_non_retirement[i] = from_non_ret
                    remaining_need -= from_non_ret
                if remaining_need > 0:
                    add_more = min(remaining_need, portfolio[i-1] - withdrawals_401k[i])
                    withdrawals_401k[i] += add_more
            real_return = (1 + investment_return) / (1 + inflation_rate) - 1
            portfolio[i] = max(0, portfolio[i-1] * (1 + real_return) - withdrawals_401k[i])
            non_retirement_savings[i] = max(0, non_retirement_savings[i-1] * (1 + real_return) - withdrawals_non_retirement[i] + excess_deposit)
    return cumulative, portfolio, withdrawals_401k, withdrawals_non_retirement, non_retirement_savings, income_taxes_paid

def create_master_table(ages, years, ss_benefit, p_values, w401k, wnr, nr, income_taxes, benefit_percentage, claim_age):
    rmds = np.zeros(len(ages))
    additional_401k = np.zeros(len(ages))
    agi = np.zeros(len(ages))
    taxable_income = np.zeros(len(ages))
    taxable_ss = np.zeros(len(ages))
    capital_gains = np.zeros(len(ages))
    recalculated_income_taxes = np.zeros(len(ages))
    for i in range(len(ages)):
        if ages[i] >= 73:
            rmd_real = calculate_rmd(ages[i], p_values[i-1] if i > 0 else initial_401k) / ((1 + inflation_rate) ** i)
            rmds[i] = rmd_real
        if i > 0 and ages[i] >= 73:
            additional_401k[i] = max(0, w401k[i] - rmds[i])
        elif i > 0:
            additional_401k[i] = w401k[i]
        capital_gains[i] = wnr[i] * non_retirement_gain_percentage
        if ss_benefit[i] > 0:
            non_ret_income = rmds[i] + additional_401k[i]
            cost_basis = wnr[i] * (1 - non_retirement_gain_percentage)
            prov_income = non_ret_income + cost_basis + (ss_benefit[i] * 0.5)
            if prov_income <= 32000:
                taxable_ss[i] = 0
            elif prov_income <= 44000:
                taxable_ss[i] = min(ss_benefit[i] * 0.5, (prov_income - 32000) * 0.5)
            else:
                taxable_ss[i] = min(ss_benefit[i] * 0.85, (6000) * 0.5 + (prov_income - 44000) * 0.85)
        agi[i] = rmds[i] + additional_401k[i] + taxable_ss[i] + capital_gains[i]
        taxable_income[i] = max(0, agi[i] - standard_deduction)
        recalculated_income_taxes[i] = calculate_federal_income_tax(taxable_income[i])
    total_income = ss_benefit + rmds + wnr + additional_401k
    total_taxes = recalculated_income_taxes
    after_tax_income = total_income - total_taxes
    excess_deposited = np.zeros(len(ages))
    for i in range(len(ages)):
        if after_tax_income[i] > target_income:
            excess_deposited[i] = after_tax_income[i] - target_income
    master_table = {
        "Age": ages,
        "Year": years,
        "% of Scheduled SS": benefit_percentage,
        "Social Security (2025$)": ss_benefit,
        "Taxable SS (2025$)": taxable_ss,
        "RMDs (2025$)": rmds,
        "Non-Retirement Withdrawals (2025$)": wnr,
        "Capital Gains (2025$)": capital_gains,
        "Additional 401k Withdrawals (2025$)": additional_401k,
        "Total Income (2025$)": total_income,
        "AGI (2025$)": agi,
        "Standard Deduction (2025$)": np.ones(len(ages)) * standard_deduction,
        "Taxable Income (2025$)": taxable_income,
        "Income Tax (2025$)": recalculated_income_taxes,
        "Total Tax (2025$)": total_taxes,
        "After-Tax Income (2025$)": after_tax_income,
        "Target After-Tax (2025$)": np.ones(len(ages)) * target_income,
        "Excess Deposited (2025$)": excess_deposited,
        "401k Balance (2025$)": p_values,
        "Non-Retirement Balance (2025$)": nr,
        "Total Portfolio (2025$)": p_values + nr
    }
    return pd.DataFrame(master_table)

# ========= FASTAPI APP SETUP =========
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/analyze")
def analyze(
    session_id: str = Query(...),
    birthdate: str = Query(...),
    age_model1: int = Query(...),
    age_model2: int = Query(...),
    fra_benefit: float = Query(...),
    inflation_rate_input: float = Query(...),
    investment_return_input: float = Query(...),
    filing_status: str = Query(...),
    initial_401k_input: float = Query(...),
    other_non_retirement_savings_input: float = Query(...),
    target_income_input: float = Query(...),
    non_retirement_gain_percentage_input: float = Query(...)
):
    # Declare all globals at the start
    global inflation_rate, investment_return, initial_401k, target_income, non_retirement_gain_percentage, tax_bracket, standard_deduction, current_age, other_non_retirement_savings

    # Set variables; note that other_retirement_income_input has been removed so we use a default.
    other_retirement_income = 0.0  # default value; remove if unused
    
    inflation_rate = inflation_rate_input
    investment_return = investment_return_input
    initial_401k = initial_401k_input
    target_income = target_income_input
    non_retirement_gain_percentage = non_retirement_gain_percentage_input
    other_non_retirement_savings = other_non_retirement_savings_input

    # Map the filing status to an approximate tax bracket.
    filing_status_map = {
         "Single": 0.22,
         "Married Filing Jointly": 0.12,
         "Married Filing Separately": 0.24,
         "Head of Household": 0.18
    }
    tax_bracket = filing_status_map.get(filing_status, 0.22)
    
    # Set standard deduction based on filing status.
    standard_deduction_map = {
         "Single": 13850,
         "Married Filing Jointly": 27700,
         "Married Filing Separately": 13850,
         "Head of Household": 20800
    }
    standard_deduction = standard_deduction_map.get(filing_status, 13850)
    
    # Compute current age from birthdate (assuming current_year = 2025)
    birth_dt = datetime.strptime(birthdate, "%Y-%m-%d")
    computed_current_age = current_year - birth_dt.year
    current_age = computed_current_age

    ages_local, years_local = get_age_ranges(computed_current_age)
    
    # Compute adjusted monthly benefits for each model.
    benefit_model1 = compute_adjusted_benefit(fra_benefit, age_model1)
    benefit_model2 = compute_adjusted_benefit(fra_benefit, age_model2)
    
    c_model1, p_model1, w401k_model1, wnr_model1, nr_model1, taxes_model1 = run_claim_strategy(
        ages_local, years_local, age_model1, benefit_model1, computed_current_age)
    c_model2, p_model2, w401k_model2, wnr_model2, nr_model2, taxes_model2 = run_claim_strategy(
        ages_local, years_local, age_model2, benefit_model2, computed_current_age)
    
    ss_model1 = np.zeros(len(ages_local))
    ss_model2 = np.zeros(len(ages_local))
    for i in range(len(ages_local)):
        if ages_local[i] >= age_model1:
            factor = benefit_reduction_factor if years_local[i] >= trust_fund_depletion_year else 1
            ss_model1[i] = benefit_model1 * factor
        if ages_local[i] >= age_model2:
            factor = benefit_reduction_factor if years_local[i] >= trust_fund_depletion_year else 1
            ss_model2[i] = benefit_model2 * factor

    benefit_percentage = np.ones(len(years_local)) * 100
    depletion_index = np.where(years_local >= trust_fund_depletion_year)[0][0] if any(years_local >= trust_fund_depletion_year) else len(years_local)
    benefit_percentage[depletion_index:] = benefit_reduction_factor * 100

    df_model1 = create_master_table(ages_local, years_local, ss_model1, p_model1, w401k_model1, wnr_model1, nr_model1, taxes_model1, benefit_percentage, age_model1)
    df_model2 = create_master_table(ages_local, years_local, ss_model2, p_model2, w401k_model2, wnr_model2, nr_model2, taxes_model2, benefit_percentage, age_model2)
    
    excel_file = f"social_security_analysis_{session_id}.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df_model1.to_excel(writer, sheet_name=f"Claim SS benefits at age {age_model1}", index=False)
        df_model2.to_excel(writer, sheet_name=f"Claim SS benefits at age {age_model2}", index=False)
        summary_data = {
            "Claiming Age": [age_model1, age_model2],
            "Monthly Benefit": [benefit_model1, benefit_model2],
            "Final 401k Balance": [p_model1[-1], p_model2[-1]],
            "Final Non-Retirement": [nr_model1[-1], nr_model2[-1]],
            "Final Portfolio Total": [p_model1[-1] + nr_model1[-1], p_model2[-1] + nr_model2[-1]],
            "Total Taxes Paid": [sum(taxes_model1), sum(taxes_model2)]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary Comparison", index=False)
    # --- Apply formatting ---
    wb = load_workbook(excel_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.alignment = Alignment(wrapText=True)
        headers = [cell.value for cell in ws[1]]
        for idx, header in enumerate(headers, start=1):
            if header and "(2025$)" in header:
                col_letter = get_column_letter(idx)
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row}"].number_format = "$#,##0"
    wb.save(excel_file)
    
    summary = {
        "Model1": df_model1.tail(1).to_dict(orient="records"),
        "Model2": df_model2.tail(1).to_dict(orient="records")
    }
    return summary

if __name__ == "__main__":
    uvicorn.run("fastapi_app:app", host="0.0.0.0", port=8000, reload=True)